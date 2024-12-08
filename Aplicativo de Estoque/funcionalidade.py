import json
import os.path
import os
import openpyxl

# Salvar os dados do estoque
arquivo_estoque = 'estoque.json'

# Carregar os dados do estoque
def carregar_estoque():
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_estoque):
        return []

    try:
        with open(arquivo_estoque, 'r') as file:
            estoque = json.load(file)
            if not isinstance(estoque, list):
                raise ValueError('Formato inválido no arquivo de estoque')
            return estoque
    except (json.JSONDecodeError, ValueError):
        return []

# Função para salvar os dodos do estoque
def salvar_estoque(estoque):
    with open(arquivo_estoque, 'w') as file:
        json.dump(estoque, file, indent=4)

# Adicionar os produtos
def adicionar_produto(nome, quantidade, preco, codigo_produto, data_cadastro):
    estoque = carregar_estoque()

    data_cadastro_str = data_cadastro.strftime('%Y-%m-%d')


    estoque.append({
                    'nome': nome,
                    'quantidade': quantidade,
                    'preco': preco,
                    'codigo_produto': codigo_produto,
                     'data_cadastro': data_cadastro_str})
    salvar_estoque(estoque)

# Função para preencher planilha do exel
def preencher_planilha(caminho_arquivo, dados_produto):
    from openpyxl import load_workbook

    # Abrir a planilha
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # Localizar a próxima linha vazia
    proxima_linha = 12
    while ws[f'J{proxima_linha}'].value is not None:
        proxima_linha += 1

    # Preencher os dados nas colunas corretas
    ws[f'J{proxima_linha}'] = dados_produto['codigo_produto']  # Código do produto
    ws[f'K{proxima_linha}'] = dados_produto['data_cadastro']  # Data de cadastro
    ws[f'L{proxima_linha}'] = dados_produto['nome']  # Nome do produto
    ws[f'M{proxima_linha}'] = dados_produto['quantidade']  # Quantidade
    ws[f'N{proxima_linha}'] = dados_produto['preco']  # Preço

    # Salvar as alterações
    wb.save(caminho_arquivo)
    wb.close()



